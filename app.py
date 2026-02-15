"""
app.py - Multi-Tenant SaaS School Exam Schedule Platform ("Luach Mivchanim").

Modules:
  1. Firebase Database Architecture (Firestore collections)
  2. Authentication & Roles (Director / Teacher / Parent-Public)
  3. Visual Scheduler (color-coded grid with Hebrew calendar)
  4. Intelligent Features (Ministry sync, conflict detection, WhatsApp share)

No hardcoded dates - all dates come from Firestore or external APIs.
"""

import hashlib
import io
import json
import re
import urllib.parse
from datetime import datetime, timedelta

import requests
import streamlit as st

from auth_manager import authenticate
from db_manager import (
    add_class_to_school,
    create_school,
    get_holidays,
    get_ministry_exam,
    get_ministry_exams,
    get_ministry_meta,
    get_permissions,
    get_schedule,
    remove_teacher_permission,
    save_ministry_exams,
    save_schedule,
    search_ministry_exams,
    set_teacher_permission,
)

# ===================================================================
# CONSTANTS (no hardcoded dates)
# ===================================================================

STYLES = {
    "bagrut":   {"bg": "#FFCDD2", "fg": "#B71C1C", "bold": True,  "label": "בגרות"},
    "magen":    {"bg": "#FFE0B2", "fg": "#E65100", "bold": True,  "label": "מגן / מתכונת"},
    "trip":     {"bg": "#C8E6C9", "fg": "#1B5E20", "bold": False, "label": "טיול / מסע"},
    "vacation": {"bg": "#BBDEFB", "fg": "#0D47A1", "bold": False, "label": "חופשה"},
    "holiday":  {"bg": "#E1BEE7", "fg": "#4A148C", "bold": False, "label": "חג / מועד"},
    "general":  {"bg": "#F5F5F5", "fg": "#424242", "bold": False, "label": "כללי"},
}

LOCKED_TYPES = {"trip"}

DAY_NAMES = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
DAY_KEYS = ["sunday", "monday", "tuesday", "wednesday", "thursday", "friday", "shabbat"]

MONTH_NAMES_HEB = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר",
}


# ===================================================================
# CSS ג€” Visual Polish
# ===================================================================

APP_CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Assistant:wght@400;600;700;800&family=Rubik:wght@400;500;600;700;800&display=swap');

/* ── Global RTL & Font ── */
html, body, .stApp {
    direction: rtl;
    font-family: 'Rubik', 'Assistant', sans-serif;
    color: #1E1E2D;
    text-rendering: optimizeLegibility;
    -webkit-font-smoothing: antialiased;
}

/* ── Hide default Streamlit chrome ── */
#MainMenu {visibility: hidden;}
header {visibility: hidden;}
footer {visibility: hidden;}
div[data-testid="stToolbar"] {display: none;}
div[data-testid="stDecoration"] {display: none;}

/* ── Main content: full width ── */
.block-container {
    padding-top: 1rem;
    padding-bottom: 0;
    max-width: 100% !important;
}

/* ── Title area (hidden when top-bar present; kept for public view) ── */
.app-title {
    text-align: center;
    padding: 1.2rem 0 0.2rem;
}
.app-title h1 {
    color: #1A237E;
    font-weight: 900;
    font-size: 2.2rem;
    font-family: 'Assistant', 'Rubik', sans-serif;
    margin: 0;
    line-height: 1.15;
    letter-spacing: 0.01em;
}
.app-title p {
    color: #9E9E9E;
    font-size: 1.05rem;
    margin: 0.15rem 0 0;
    font-weight: 500;
}
/* Hide duplicate title inside tabs (top-bar already shows it) */
div[data-testid="stTabs"] .app-title { display: none; }

/* ── Schedule table ── */
.sched-table {
    width: 100%;
    border-collapse: collapse;
    direction: rtl;
    table-layout: fixed;
}
.sched-table th, .sched-table td {
    border: 1px solid #DEE2E6;
    text-align: center;
    vertical-align: top;
}

/* ── Calendar header row ── */
.cal-hdr, .sched-table th.cal-hdr {
    background: linear-gradient(135deg, #1A237E, #283593);
    color: #fff;
    font-weight: 700;
    font-size: 0.82em;
    text-align: center;
    padding: 10px 4px;
    border-radius: 0;
}

/* ── Legend chips ── */
.legend-row {
    display: flex; flex-wrap: wrap; gap: 6px;
    justify-content: center; margin: 8px 0 14px;
}
.legend-chip {
    padding: 4px 14px; border-radius: 20px;
    font-size: 0.78em; font-weight: 500;
    box-shadow: 0 1px 2px rgba(0,0,0,0.08);
}
.cal-parasha {
    background: #FFF8E1; color: #F57F17; font-weight: 700;
    padding: 2px 7px; border-radius: 10px; font-size: 0.72em;
}

/* ── Cell popover trigger (small + button) ── */
.stColumn > div { gap: 2px; }
div[data-testid="stPopover"] button {
    font-size: 0.75em !important; padding: 0 4px !important;
    min-height: 0 !important; height: 20px !important;
    line-height: 20px !important; border-radius: 4px !important;
    background: #E8EAF6 !important; border: 1px solid #C5CAE9 !important;
    color: #3949AB !important; cursor: pointer !important;
    width: 100% !important;
}
div[data-testid="stPopover"] button:hover { background: #C5CAE9 !important; }

/* ── Ministry card ── */
.ministry-card {
    background: #E8F5E9; border: 1px solid #A5D6A7; border-radius: 8px;
    padding: 10px 14px; margin: 6px 0;
}

/* ------------------------------------------------------------------
   LAYOUT DESIGN SYSTEM (no sidebar - all inline)
   ------------------------------------------------------------------ */
:root {
    --ds-surface: #FFFFFF;
    --ds-bg-subtle: #F8F9FC;
    --ds-border: #E2E5EF;
    --ds-border-subtle: #ECEEF5;
    --ds-primary: #1A237E;
    --ds-primary-light: #E8EAF6;
    --ds-primary-medium: #C5CAE9;
    --ds-text: #1E1E2D;
    --ds-text-muted: #6B7294;
    --ds-primary-strong: #2434A6;
    --ds-radius: 10px;
    --ds-radius-sm: 6px;
    --ds-shadow: 0 1px 3px rgba(26,35,126,0.06), 0 1px 2px rgba(0,0,0,0.04);
    --ds-shadow-hover: 0 4px 12px rgba(26,35,126,0.10);
    --ds-transition: all 0.15s ease;
}

/* ── Hide sidebar toggle (not used) ── */
section[data-testid="stSidebar"] { display: none !important; }
button[data-testid="stSidebarCollapsedControl"] { display: none !important; }

/* ── Top bar ── */
.top-bar {
    display: flex; align-items: center; justify-content: space-between;
    background: linear-gradient(135deg, #1A237E 0%, #283593 100%);
    border-radius: 14px; padding: 14px 24px;
    box-shadow: 0 4px 20px rgba(26,35,126,0.18);
    margin-bottom: 10px; gap: 16px; flex-wrap: wrap;
    color: #fff;
}
.top-bar-user {
    display: flex; align-items: center; gap: 10px; flex-shrink: 0;
}
.top-bar-avatar {
    width: 40px; height: 40px; border-radius: 50%;
    background: rgba(255,255,255,0.2);
    backdrop-filter: blur(4px); -webkit-backdrop-filter: blur(4px);
    border: 2px solid rgba(255,255,255,0.35);
    color: #fff; display: flex; align-items: center; justify-content: center;
    font-weight: 800; font-size: 1rem; flex-shrink: 0;
}
.top-bar-name { font-weight: 700; font-size: 0.88rem; color: #fff; }
.top-bar-email { font-size: 0.72rem; color: rgba(255,255,255,0.7); }
.top-bar-role {
    font-size: 0.62rem; font-weight: 700; padding: 2px 10px;
    border-radius: 10px; display: inline-block; margin-top: 2px;
}
.top-bar-role.director { background: rgba(255,255,255,0.2); color: #fff; }
.top-bar-role.teacher  { background: rgba(255,255,255,0.15); color: #C8E6C9; }
.top-bar-title {
    text-align: center; flex: 1;
}
.top-bar-title h2 {
    font-weight: 900; font-size: 1.5rem; color: #fff;
    font-family: 'Assistant', 'Rubik', sans-serif;
    margin: 0; line-height: 1.15; letter-spacing: 0.01em;
}
.top-bar-title p {
    font-size: 0.82rem; color: rgba(255,255,255,0.7);
    margin: 2px 0 0; font-weight: 500;
}

/* ── Controls row (class + date) ── */
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .ctrl-row-marker) {
    background: var(--ds-surface);
    border: 1px solid var(--ds-border);
    border-radius: var(--ds-radius);
    padding: 8px 12px 4px;
    box-shadow: var(--ds-shadow);
    margin-bottom: 6px;
    position: sticky;
    top: 10px;
    z-index: 30;
    backdrop-filter: blur(6px);
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .ctrl-row-marker) label {
    font-size: 0.74rem; font-weight: 600; color: var(--ds-text-muted);
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .ctrl-row-marker) .stSelectbox > div > div {
    border-radius: 6px; border-color: var(--ds-border);
    font-weight: 600; font-size: 0.83rem; min-height: 34px;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .ctrl-row-marker) .stDateInput > div > div > input {
    border-radius: 6px; font-weight: 500; font-size: 0.8rem; padding: 5px 8px;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .ctrl-row-marker) .stSelectbox,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .ctrl-row-marker) .stDateInput {
    max-width: 280px;
}
.ctrl-row-marker { display: none; }

div[data-testid="stRadio"] [role="radiogroup"] {
    gap: 0.45rem !important;
}
div[data-testid="stRadio"] label p {
    white-space: nowrap !important;
    word-break: keep-all !important;
    line-height: 1.15 !important;
}

/* ── Export toolbar ── */
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .export-bar-marker) {
    background: var(--ds-bg-subtle);
    border: 1px solid var(--ds-border-subtle);
    border-radius: 8px;
    padding: 6px 10px 2px;
    margin-bottom: 6px;
}
div[data-testid="stVerticalBlock"]:has(.export-bar-marker) .stDownloadButton > button,
div[data-testid="stVerticalBlock"]:has(.export-bar-marker) .stButton > button {
    font-size: 0.76rem !important; padding: 5px 10px !important;
    border-radius: 6px !important; font-weight: 600 !important;
    min-height: 34px !important;
}
div[data-testid="stVerticalBlock"]:has(.export-bar-marker) div[data-testid="stPopover"] button {
    font-size: 0.76rem !important; padding: 5px 10px !important;
    height: 34px !important; min-height: 0 !important;
    line-height: normal !important; border-radius: 6px !important;
    background: var(--ds-surface) !important; border: 1px solid var(--ds-border) !important;
    color: var(--ds-text) !important; font-weight: 600 !important;
}
div[data-testid="stVerticalBlock"]:has(.export-bar-marker) > div[data-testid="stHorizontalBlock"] > div:nth-child(1) div[data-testid="stPopover"] button {
    background: #EEF2FF !important; border-color: #A5B4FC !important; color: #3730A3 !important;
}
div[data-testid="stVerticalBlock"]:has(.export-bar-marker) > div[data-testid="stHorizontalBlock"] > div:nth-child(2) .stDownloadButton > button {
    background: #ECFDF3 !important; border-color: #86EFAC !important; color: #166534 !important;
}
div[data-testid="stVerticalBlock"]:has(.export-bar-marker) > div[data-testid="stHorizontalBlock"] > div:nth-child(3) .stDownloadButton > button {
    background: #FFF7ED !important; border-color: #FDBA74 !important; color: #9A3412 !important;
}

/* ── Tabs ── */
div[data-testid="stTabs"] button[data-baseweb="tab"] {
    font-family: 'Assistant', 'Rubik', sans-serif;
    font-weight: 700; font-size: 0.9rem;
    padding: 8px 22px;
    color: var(--ds-text-muted);
    border-bottom: 3px solid transparent;
    transition: var(--ds-transition);
}
div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"] {
    color: var(--ds-primary);
    border-bottom-color: var(--ds-primary);
}
div[data-testid="stTabs"] button[data-baseweb="tab"]:hover {
    color: var(--ds-primary);
}

/* ── Section headers inside tabs ── */
.sb-section-hdr {
    display: flex; align-items: center; gap: 8px;
    font-family: 'Assistant', 'Rubik', sans-serif;
    font-size: 0.94rem; font-weight: 800; color: var(--ds-primary);
    letter-spacing: 0.02em; padding: 0 0 8px; margin: 18px 0 12px;
    border-bottom: 2px solid var(--ds-primary-light);
}
.sb-section-hdr .sb-icon { font-size: 1.05rem; line-height: 1; }

/* ── Admin expanders (card-style) ── */
.admin-panel details[data-testid="stExpander"] {
    background: var(--ds-surface);
    border: 1px solid var(--ds-border);
    border-radius: var(--ds-radius) !important;
    box-shadow: var(--ds-shadow);
    margin-bottom: 8px; overflow: hidden;
}
.admin-panel details[data-testid="stExpander"] summary {
    font-weight: 600; font-size: 0.88rem;
    color: var(--ds-text); padding: 12px 16px;
}

/* ── Staff cards ── */
.staff-card {
    background: var(--ds-surface); border: 1px solid var(--ds-border);
    border-radius: var(--ds-radius); padding: 10px 14px;
    margin-bottom: 6px; box-shadow: var(--ds-shadow);
    transition: var(--ds-transition);
}
.staff-card:hover {
    border-color: var(--ds-primary-medium);
    box-shadow: var(--ds-shadow-hover);
}
.staff-email {
    font-weight: 600; font-size: 0.88rem; color: var(--ds-text);
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.staff-meta { font-size: 0.78rem; color: var(--ds-text-muted); margin-top: 3px; }
.class-chip {
    display: inline-block; background: var(--ds-primary-light);
    color: var(--ds-primary); font-size: 0.72rem; font-weight: 600;
    padding: 2px 10px; border-radius: 8px; margin: 2px 0 2px 4px;
}
.staff-empty {
    text-align: center; padding: 24px 10px;
    color: var(--ds-text-muted); font-size: 0.88rem;
    background: var(--ds-bg-subtle); border: 1px dashed var(--ds-border);
    border-radius: var(--ds-radius);
}

/* ── Export area ── */
.export-grid {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 12px; margin-top: 12px;
}
.sb-export-link {
    display: block; text-align: center; padding: 12px;
    border-radius: var(--ds-radius-sm); text-decoration: none;
    font-weight: 600; font-size: 0.9rem;
    transition: var(--ds-transition); box-shadow: var(--ds-shadow);
}
.sb-export-link:hover {
    box-shadow: var(--ds-shadow-hover); transform: translateY(-1px);
}

/* ── Nav button bar ── */
.nav-bar {
    display: flex; gap: 8px; flex-wrap: wrap;
    justify-content: center; margin: 14px 0 10px;
    padding: 8px 12px;
    background: var(--ds-bg-subtle);
    border: 1px solid var(--ds-border);
    border-radius: var(--ds-radius);
}
.nav-btn {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 8px 18px; border-radius: 8px;
    font-family: 'Heebo', sans-serif; font-weight: 600; font-size: 0.88rem;
    border: 1.5px solid var(--ds-border); background: var(--ds-surface);
    color: var(--ds-text-muted); cursor: pointer;
    transition: var(--ds-transition); text-decoration: none;
    box-shadow: var(--ds-shadow);
}
.nav-btn:hover {
    border-color: var(--ds-primary-medium);
    color: var(--ds-primary); background: var(--ds-primary-light);
    box-shadow: var(--ds-shadow-hover); transform: translateY(-1px);
}
.nav-btn.active {
    background: var(--ds-primary); color: #fff;
    border-color: var(--ds-primary); box-shadow: var(--ds-shadow-hover);
}
.nav-btn .nav-icon { font-size: 1.1rem; line-height: 1; }

/* ── Unified button system ── */
.stButton > button, .stDownloadButton > button {
    font-family: 'Rubik', 'Assistant', sans-serif !important;
    font-weight: 700 !important;
    font-size: 0.86rem !important;
    letter-spacing: 0.01em !important;
    border-radius: 10px !important;
    border: 1px solid #D8DDF0 !important;
    box-shadow: 0 2px 6px rgba(31, 41, 55, 0.08) !important;
    transition: transform 0.14s ease, box-shadow 0.2s ease, border-color 0.2s ease, background 0.2s ease !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 8px 16px rgba(31, 41, 55, 0.12) !important;
}
.stButton > button[kind="primary"], .stDownloadButton > button[kind="primary"] {
    background: linear-gradient(135deg, var(--ds-primary-strong), var(--ds-primary)) !important;
    color: #fff !important;
    border-color: transparent !important;
    box-shadow: 0 10px 20px rgba(26, 35, 126, 0.28) !important;
}
.stButton > button[kind="primary"]:hover, .stDownloadButton > button[kind="primary"]:hover {
    filter: brightness(1.03);
}
/* Side action buttons */
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) .stButton > button {
    font-size: 0.84rem !important;
    padding: 8px 12px !important;
    min-height: 40px !important;
    border-radius: 10px !important;
    border: 1.5px solid #D8DDF0 !important;
    background: var(--ds-surface) !important;
    color: var(--ds-text) !important;
    box-shadow: var(--ds-shadow) !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    gap: 8px !important;
    text-align: right !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) .stButton > button:hover {
    border-color: #B4BEE6 !important;
    box-shadow: var(--ds-shadow-hover) !important;
    filter: brightness(0.99);
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"]:nth-of-type(2) .stButton > button {
    background: #FFF1F2 !important; border-color: #FDA4AF !important; color: #9F1239 !important; border-right: 4px solid #E11D48 !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"]:nth-of-type(3) .stButton > button {
    background: #EEF2FF !important; border-color: #A5B4FC !important; color: #3730A3 !important; border-right: 4px solid #4F46E5 !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"]:nth-of-type(4) .stButton > button {
    background: #ECFDF3 !important; border-color: #86EFAC !important; color: #166534 !important; border-right: 4px solid #16A34A !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"]:nth-of-type(5) .stButton > button {
    background: #F0F9FF !important; border-color: #7DD3FC !important; color: #075985 !important; border-right: 4px solid #0284C7 !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"]:nth-of-type(6) .stButton > button {
    background: #FFF7ED !important; border-color: #FDBA74 !important; color: #9A3412 !important; border-right: 4px solid #EA580C !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"]:nth-of-type(7) .stButton > button {
    background: #F5F3FF !important; border-color: #C4B5FD !important; color: #5B21B6 !important; border-right: 4px solid #7C3AED !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-actions-marker) > div[data-testid="element-container"] .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, var(--ds-primary-strong), var(--ds-primary)) !important;
    color: #fff !important;
    border-color: transparent !important;
    box-shadow: 0 12px 22px rgba(26, 35, 126, 0.3) !important;
}
.side-actions-title {
    font-family: 'Assistant', 'Rubik', sans-serif;
    font-size: 0.82rem;
    font-weight: 800;
    color: var(--ds-primary);
    margin: 2px 2px 8px;
}
.side-actions-marker { display: none; }

/* Side panel container (stable marker-based scope) */
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) {
    background: var(--ds-surface);
    border: 1px solid var(--ds-border);
    border-radius: var(--ds-radius);
    padding: 10px 12px 8px;
    box-shadow: var(--ds-shadow);
    margin-top: 8px;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .side-panel-heading {
    font-size: 0.88rem;
    font-family: 'Assistant', 'Rubik', sans-serif;
    font-weight: 800;
    color: var(--ds-primary);
    margin: 0 0 10px;
    padding-bottom: 6px;
    border-bottom: 2px solid var(--ds-primary-light);
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stTextInput > div > div > input,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stSelectbox > div > div,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stMultiSelect > div > div {
    font-size: 0.84rem !important;
    min-height: 36px !important;
    border-radius: 6px !important;
}
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stTextInput label,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stSelectbox label,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stMultiSelect label,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stDateInput label,
div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div[data-testid="stMarkdown"] .side-panel-marker) .stNumberInput label {
    font-size: 0.76rem !important;
    font-weight: 600 !important;
    color: var(--ds-text-muted) !important;
}
.side-panel-marker { display: none; }

label, .stCaption, .stMarkdown p, .stTextInput input, .stSelectbox, .stMultiSelect {
    font-family: 'Rubik', 'Assistant', sans-serif !important;
}

/* ── Logout button ── */
.logout-btn .stButton > button {
    font-size: 0.78rem !important;
    padding: 6px 12px !important;
    min-height: 34px !important;
    border-radius: 8px !important;
    background: transparent !important;
    border: 1.5px solid rgba(255,255,255,0.25) !important;
    color: #E57373 !important;
    font-weight: 600 !important;
}
.logout-btn .stButton > button:hover {
    background: rgba(229,115,115,0.1) !important;
    border-color: #E57373 !important;
}
</style>"""


# ===================================================================
# HELPERS
# ===================================================================

def get_day_date(start_date_str: str, day_index: int) -> str:
    try:
        sd = datetime.strptime(start_date_str, "%Y-%m-%d")
        d = sd + timedelta(days=day_index)
        return d.strftime("%d/%m")
    except Exception:
        return ""


def get_full_date(start_date_str: str, day_index: int):
    try:
        sd = datetime.strptime(start_date_str, "%Y-%m-%d")
        return sd + timedelta(days=day_index)
    except Exception:
        return None


def date_to_week_day(weeks: list, target_date) -> tuple | None:
    for wi, wk in enumerate(weeks):
        try:
            sd = datetime.strptime(wk["start_date"], "%Y-%m-%d")
        except Exception:
            continue
        for di, dk in enumerate(DAY_KEYS):
            d = sd + timedelta(days=di)
            if d.date() == target_date.date():
                return wi, dk
    return None


def check_conflicts_on_date(weeks: list, wi: int, dk: str, cls: str) -> list[str]:
    events = weeks[wi]["days"].get(dk, [])
    return [
        ev["text"] for ev in events
        if ev.get("type") in LOCKED_TYPES and ev.get("class") in (cls, "all")
    ]


def fetch_parasha_from_api(start_year: int, end_year: int) -> dict:
    """Fetch Parashat HaShavua from Hebcal API. Returns {sunday_date: hebrew_name}."""
    parasha_map = {}
    for year in range(start_year, end_year + 1):
        try:
            url = (
                f"https://www.hebcal.com/hebcal?v=1&cfg=json&s=on"
                f"&year={year}&month=x&geo=geoname&geonameid=281184"
            )
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                for item in data.get("items", []):
                    if item.get("category") == "parashat":
                        d = item.get("date", "")
                        title = item.get("hebrew", item.get("title", ""))
                        if d and title:
                            shabbat = datetime.strptime(d, "%Y-%m-%d")
                            sunday = shabbat - timedelta(days=6)
                            parasha_map[sunday.strftime("%Y-%m-%d")] = title
        except Exception:
            pass
    return parasha_map


def refresh_ministry_db_from_web(season: str = "summer") -> int:
    """Download Ministry of Education exam schedule Excel and store in Firestore."""
    from openpyxl import load_workbook

    current_year = datetime.now().year
    if season == "summer":
        url = f"https://meyda.education.gov.il/files/Exams/HoursSumExams{current_year}.xlsx"
        moed_label = f'מועד קיץ {current_year}'
    else:
        url = f"https://meyda.education.gov.il/files/Exams/LuachWinExams{current_year}HOURS.xlsx"
        moed_label = f'מועד חורף {current_year}'

    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb.active

    exams = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        date_val, code, name = row[0], row[1], row[2]
        start_time = row[3] if len(row) > 3 else None
        end_time = row[4] if len(row) > 4 else None
        if not date_val or not code or not name:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)
        st_str = ""
        et_str = ""
        if start_time and hasattr(start_time, "strftime"):
            st_str = start_time.strftime("%H:%M")
        elif start_time:
            st_str = str(start_time)
        if end_time and hasattr(end_time, "strftime"):
            et_str = end_time.strftime("%H:%M")
        elif end_time:
            et_str = str(end_time)
        exams.append({
            "code": str(int(code)) if isinstance(code, (int, float)) else str(code),
            "name": str(name).strip(),
            "date": date_str,
            "start_time": st_str,
            "end_time": et_str,
        })

    save_ministry_exams(exams, moed=moed_label, source="משרד החינוך - אגף בחינות")
    return len(exams)


def generate_new_year(start_year: int) -> dict:
    """Generate a new academic year schedule structure. No hardcoded dates."""
    sep1 = datetime(start_year, 9, 1)
    days_since_sunday = (sep1.weekday() + 1) % 7
    first_sunday = sep1 - timedelta(days=days_since_sunday)
    aug31 = datetime(start_year + 1, 8, 31)

    weeks = []
    current = first_sunday
    while current <= aug31:
        end = current + timedelta(days=6)
        if current.month == end.month:
            dr = f"{current.day}-{end.day}.{current.month}"
        else:
            dr = f"{current.day}.{current.month}-{end.day}.{end.month}"
        weeks.append({
            "date_range": dr,
            "start_date": current.strftime("%Y-%m-%d"),
            "days": {dk: [] for dk in DAY_KEYS},
        })
        current += timedelta(days=7)

    heb_year_num = start_year + 3761
    heb_letters = {
        5784: 'תשפ"ד', 5785: 'תשפ"ה', 5786: 'תשפ"ו', 5787: 'תשפ"ז',
        5788: 'תשפ"ח', 5789: 'תשפ"ט', 5790: 'תש"צ',
    }
    year_label = heb_letters.get(heb_year_num, f"תשפ {heb_year_num - 5000}")

    new_data = {
        "classes": [],
        "year": year_label,
        "weeks": weeks,
        "parashat_hashavua": {},
    }

    for yr_key in [str(start_year), str(start_year + 1)]:
        holidays_data = get_holidays(yr_key)
        if not holidays_data:
            continue
        if "label" in holidays_data and yr_key == str(start_year):
            new_data["year"] = holidays_data["label"]
        for h in holidays_data.get("holidays", []):
            try:
                hdate = datetime.strptime(h["date"], "%Y-%m-%d")
            except Exception:
                continue
            loc = date_to_week_day(weeks, hdate)
            if loc is None:
                continue
            wi, dk = loc
            weeks[wi]["days"][dk].append({
                "text": h["text"], "type": h.get("type", "holiday"), "class": "all",
            })
        for v in holidays_data.get("school_vacations", []):
            try:
                vs = datetime.strptime(v["start"], "%Y-%m-%d")
                ve = datetime.strptime(v["end"], "%Y-%m-%d")
            except Exception:
                continue
            d = vs
            while d <= ve:
                loc = date_to_week_day(weeks, d)
                if loc:
                    wi, dk = loc
                    existing_texts = [e["text"] for e in weeks[wi]["days"][dk]]
                    if v["text"] not in existing_texts:
                        weeks[wi]["days"][dk].append({
                            "text": v["text"], "type": "vacation", "class": "all",
                        })
                d += timedelta(days=1)

    try:
        parasha = fetch_parasha_from_api(start_year, start_year + 1)
        new_data["parashat_hashavua"] = parasha
    except Exception:
        new_data["parashat_hashavua"] = {}

    return new_data


# ===================================================================
# IMPORT EXAM TO SCHEDULE
# ===================================================================

def _normalize_exam_time(value) -> str:
    """Normalize exam time values to HH:MM when possible."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%H:%M")
        except Exception:
            pass
    raw = str(value).strip()
    if not raw:
        return ""
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", raw)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return raw


def _build_bagrut_label(exam: dict) -> str:
    """Create a display label for bagrut events including times when available."""
    name = exam.get("name", "")
    code = exam.get("code", "")
    start_time = _normalize_exam_time(exam.get("start_time"))
    end_time = _normalize_exam_time(exam.get("end_time"))

    base = f"בגרות {name} ({code})".strip()
    if start_time and end_time:
        return f"{base} {start_time}-{end_time}"
    if start_time:
        return f"{base} {start_time}"
    return base

def import_exam_to_schedule(data: dict, exam: dict, cls: str) -> tuple[bool, str]:
    """Add a ministry exam to the schedule. Returns (success, message)."""
    try:
        target = datetime.strptime(exam["date"], "%Y-%m-%d")
    except Exception:
        return False, "תאריך לא תקין"

    loc = date_to_week_day(data["weeks"], target)
    if loc is None:
        return False, "התאריך לא נמצא בטווח השבועות של הלוח"

    wi, dk = loc
    conflict_msg = ""
    conflicts = check_conflicts_on_date(data["weeks"], wi, dk, cls)
    if conflicts:
        conflict_msg = f"שים לב: התאריך מתנגש עם אירוע קיים! ({', '.join(conflicts)})"

    label = _build_bagrut_label(exam)
    new_event = {
        "text": label,
        "type": "bagrut",
        "class": cls,
        "exam_code": exam["code"],
        "start_time": _normalize_exam_time(exam.get("start_time")),
        "end_time": _normalize_exam_time(exam.get("end_time")),
    }

    cell = data["weeks"][wi]["days"].get(dk, [])
    for ev in cell:
        if ev.get("exam_code") == exam["code"] and ev.get("class") == cls:
            return False, "הבגרות כבר קיימת בלוח בתאריך זה"
    cell.append(new_event)
    data["weeks"][wi]["days"][dk] = cell
    return True, conflict_msg


def resync_dates_with_ministry(data: dict, cls: str) -> list[dict]:
    """Re-check all bagrut events against Firestore ministry data. Move if dates changed."""
    all_exams = get_ministry_exams()
    ministry_lookup = {ex["code"]: ex for ex in all_exams if ex.get("code") != "_metadata"}
    changes = []

    for wi, wk in enumerate(data["weeks"]):
        try:
            sd = datetime.strptime(wk["start_date"], "%Y-%m-%d")
        except Exception:
            continue
        for di, dk in enumerate(DAY_KEYS):
            cell = wk["days"].get(dk, [])
            current_date = sd + timedelta(days=di)
            for ev in list(cell):
                code = ev.get("exam_code")
                if not code or ev.get("class") not in (cls, "all"):
                    continue
                if code not in ministry_lookup:
                    continue
                official = ministry_lookup[code]
                # Keep event text and stored hours aligned with latest ministry data.
                ev["text"] = _build_bagrut_label(official)
                ev["start_time"] = _normalize_exam_time(official.get("start_time"))
                ev["end_time"] = _normalize_exam_time(official.get("end_time"))
                try:
                    official_date = datetime.strptime(official["date"], "%Y-%m-%d")
                except Exception:
                    continue
                if current_date.date() == official_date.date():
                    continue
                new_loc = date_to_week_day(data["weeks"], official_date)
                if new_loc is None:
                    changes.append({
                        "code": code, "name": official["name"],
                        "old_date": current_date.strftime("%d/%m/%Y"),
                        "new_date": official_date.strftime("%d/%m/%Y"),
                        "conflict": "התאריך החדש מחוץ לטווח הלוח",
                    })
                    continue
                new_wi, new_dk = new_loc
                conflict_list = check_conflicts_on_date(data["weeks"], new_wi, new_dk, cls)
                conflict_msg = f"התנגשות עם: {', '.join(conflict_list)}" if conflict_list else ""
                cell.remove(ev)
                new_cell = data["weeks"][new_wi]["days"].get(new_dk, [])
                new_cell.append(ev)
                data["weeks"][new_wi]["days"][new_dk] = new_cell
                changes.append({
                    "code": code, "name": official["name"],
                    "old_date": current_date.strftime("%d/%m/%Y"),
                    "new_date": official_date.strftime("%d/%m/%Y"),
                    "conflict": conflict_msg,
                })
    return changes


# ===================================================================
# HTML RENDERING
# ===================================================================

def chip_html(ev: dict) -> str:
    s = STYLES.get(ev.get("type", "general"), STYLES["general"])
    w = "700" if s["bold"] else "400"
    return (
        f'<span style="background:{s["bg"]};color:{s["fg"]};font-weight:{w};'
        f'padding:2px 8px;border-radius:10px;font-size:0.78em;display:inline-block;'
        f'margin:1px 0;line-height:1.4;">{ev["text"]}</span>'
    )


def cell_html(date_str: str, chips_html: str, even: bool = False) -> str:
    bg = "#F8F9FA" if even else "#FFFFFF"
    return (
        f'<div style="background:{bg};border:1px solid #DEE2E6;border-radius:6px;'
        f'padding:4px 3px;min-height:72px;text-align:center;display:flex;'
        f'flex-direction:column;align-items:center;justify-content:flex-start;gap:2px;'
        f'overflow:visible;">'
        f'<div style="color:#90A4AE;font-size:0.7em;font-weight:500;flex-shrink:0;">{date_str}</div>'
        f'{chips_html}</div>'
    )


def exam_card_html(exam: dict) -> str:
    try:
        d = datetime.strptime(exam["date"], "%Y-%m-%d")
        date_display = d.strftime("%d/%m/%Y")
    except Exception:
        date_display = exam.get("date", "")
    details = f'סמל: {exam["code"]} | תאריך: {date_display}'
    if exam.get("start_time"):
        details += f' | {exam["start_time"]}-{exam.get("end_time", "")}'
    return (
        f'<div class="ministry-card"><b>{exam["name"]}</b><br>{details}</div>'
    )


# ===================================================================
# EXPORT: EXCEL
# ===================================================================

def to_excel(data: dict, cls: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cols = ["שבוע"] + DAY_NAMES
    pm = data.get("parashat_hashavua", {})
    wb = Workbook()
    ws = wb.active
    ws.title = cls
    ws.sheet_view.rightToLeft = True
    hf = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    hfn = Font(color="FFFFFF", bold=True, size=11)
    b = Border(*(Side(style="thin", color="B0BEC5") for _ in range(4)))
    for ci, cn in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.fill = hf
        c.font = hfn
        c.border = b
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, wk in enumerate(data["weeks"], 2):
        p = pm.get(wk["start_date"], "")
        c = ws.cell(row=ri, column=1, value=wk["date_range"])
        c.font = Font(bold=True, size=10)
        c.border = b
        c.alignment = Alignment(horizontal="center", vertical="center")
        for di, dk in enumerate(DAY_KEYS):
            evs = [e for e in wk["days"].get(dk, []) if e.get("class") in (cls, "all")]
            tx = [e["text"] for e in evs]
            if dk == "shabbat" and p:
                tx.append(f"פרשת {p}")
            cell = ws.cell(row=ri, column=di + 2, value="\n".join(tx))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = b
            if evs:
                pr = ["bagrut", "magen", "trip", "vacation", "holiday"]
                dm = next((x for x in pr if any(e["type"] == x for e in evs)), "general")
                st2 = STYLES[dm]
                cell.fill = PatternFill(
                    start_color=st2["bg"].lstrip("#"),
                    end_color=st2["bg"].lstrip("#"),
                    fill_type="solid",
                )
                cell.font = Font(color=st2["fg"].lstrip("#"), bold=st2["bold"], size=10)
    ws.column_dimensions["A"].width = 14
    for ch in "BCDEFGH":
        ws.column_dimensions[ch].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===================================================================
# EXPORT: PNG / HTML
# ===================================================================

def _build_schedule_html(data: dict, cls: str, filtered_weeks: list) -> str:
    """Build a self-contained HTML string of the schedule table for rendering."""
    pm = data.get("parashat_hashavua", {})
    html_parts = [
        '<html><head><meta charset="utf-8"><style>'
        '@import url("https://fonts.googleapis.com/css2?family=Heebo:wght@400;700&display=swap");'
        'body{direction:rtl;font-family:"Heebo",sans-serif;margin:16px;background:#fff;}'
        'h3{text-align:center;color:#1A237E;margin:8px 0 12px;}'
        'table{border-collapse:collapse;width:100%;direction:rtl;}'
        'th{background:#1A237E;color:#fff;font-weight:700;padding:10px 8px;font-size:13px;border:1px solid #B0BEC5;}'
        'td{padding:6px 5px;font-size:12px;border:1px solid #DEE2E6;text-align:center;vertical-align:top;min-width:110px;}'
        '.date-label{color:#90A4AE;font-size:0.8em;}'
        '.parasha{color:#F57F17;font-weight:700;}'
        '</style></head><body>'
        f'<h3>לוח שנה {data.get("year","")} - {cls}</h3>'
        '<table><thead><tr>'
    ]
    for dn in DAY_NAMES:
        html_parts.append(f'<th>{dn}</th>')
    html_parts.append('</tr></thead><tbody>')

    for wi, wk in filtered_weeks:
        parasha = pm.get(wk["start_date"], "")
        html_parts.append('<tr>')
        for di, dk in enumerate(DAY_KEYS):
            day_date = get_day_date(wk.get("start_date", ""), di)
            evs = [e for e in wk["days"].get(dk, []) if e.get("class") in (cls, "all")]
            bg_color = "#FFFFFF" if wi % 2 else "#F8F9FA"
            if evs:
                pr = ["bagrut", "magen", "trip", "vacation", "holiday"]
                dm = next((x for x in pr if any(e["type"] == x for e in evs)), "general")
                bg_color = STYLES[dm]["bg"]
            cell_texts = [f'<small class="date-label">{day_date}</small>']
            for ev in evs:
                s = STYLES.get(ev["type"], STYLES["general"])
                cell_texts.append(
                    f'<span style="color:{s["fg"]};font-weight:{"700" if s["bold"] else "400"};">'
                    f'{ev["text"]}</span>'
                )
            if dk == "shabbat" and parasha:
                cell_texts.append(f'<span class="parasha">{parasha}</span>')
            html_parts.append(f'<td style="background:{bg_color};">{"<br>".join(cell_texts)}</td>')
        html_parts.append('</tr>')

    html_parts.append('</tbody></table></body></html>')
    return "".join(html_parts)


def schedule_to_png(data: dict, cls: str, filtered_weeks: list):
    """Render the schedule as a high-resolution PNG image using Playwright."""
    full_html = _build_schedule_html(data, cls, filtered_weeks)

    try:
        import asyncio
        import subprocess
        import sys

        # Windows requires ProactorEventLoop for subprocess support.
        # Streamlit's event loop doesn't set this, so Playwright crashes
        # with NotImplementedError on asyncio.create_subprocess_exec.
        if sys.platform == "win32":
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

        from playwright.sync_api import sync_playwright

        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                browser.close()
        except Exception:
            subprocess.run(
                ["playwright", "install", "chromium"],
                check=True,
                capture_output=True,
                timeout=120,
            )

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(
                viewport={"width": 1200, "height": 800},
                device_scale_factor=2,
            )
            page.set_content(full_html, wait_until="networkidle")
            page.wait_for_timeout(1500)
            png_bytes = page.screenshot(full_page=True, type="png")
            browser.close()
        return png_bytes
    except Exception:
        return full_html.encode("utf-8")


# ===================================================================
# WHATSAPP SHARE
# ===================================================================

def build_whatsapp_text(data: dict, cls: str, filtered_weeks: list) -> str:
    pm = data.get("parashat_hashavua", {})
    lines = [f"*לוח שנה {data.get('year', '')} - {cls}*\n"]
    for wi, wk in filtered_weeks:
        parasha = pm.get(wk["start_date"], "")
        week_has_events = False
        week_lines = [f"\n*שבוע {wk['date_range']}*"]
        if parasha:
            week_lines.append(f"  פרשת {parasha}")
        for di, dk in enumerate(DAY_KEYS):
            day_date = get_day_date(wk.get("start_date", ""), di)
            evs = [e for e in wk["days"].get(dk, []) if e.get("class") in (cls, "all")]
            if evs:
                week_has_events = True
                day_label = DAY_NAMES[di]
                for ev in evs:
                    week_lines.append(f"  {day_label} {day_date}: {ev['text']}")
        if week_has_events:
            lines.extend(week_lines)
    return "\n".join(lines)


def _export_cache_key(data: dict, cls: str, filtered_weeks: list) -> str:
    """Stable export fingerprint so heavy generators rerun only on real data changes."""
    payload = {
        "year": data.get("year", ""),
        "class": cls,
        "weeks": [],
    }
    for _, wk in filtered_weeks:
        week_payload = {"start_date": wk.get("start_date", ""), "days": {}}
        for dk in DAY_KEYS:
            events = [e for e in wk["days"].get(dk, []) if e.get("class") in (cls, "all")]
            if events:
                week_payload["days"][dk] = events
        payload["weeks"].append(week_payload)
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


# ===================================================================
# PAGES
# ===================================================================

def page_create_school(auth_info: dict):
    """Page shown when a director has no school yet - create one."""
    st.markdown(
        '<div class="app-title"><h1>יצירת מוסד חדש</h1>'
        '<p>צור את מוסד הלימודים שלך כדי להתחיל</p></div>',
        unsafe_allow_html=True,
    )
    with st.form("create_school_form"):
        school_name = st.text_input("שם המוסד", placeholder="תיכון הדמוקרטי")
        school_id = st.text_input("מזהה (אנגלית)", placeholder="my-school")
        classes_input = st.text_input("כיתות (פסיק)", value="יא 1, יא 2, יא 3")
        submitted = st.form_submit_button("צור מוסד", type="primary", use_container_width=True)
        if submitted and school_name.strip() and school_id.strip():
            classes = [c.strip() for c in classes_input.split(",") if c.strip()]
            if not classes:
                classes = ["יא 1", "יא 2", "יא 3"]
            try:
                create_school(school_id.strip(), auth_info["email"], school_name.strip(), classes)
                current_year = datetime.now().year
                month_now = datetime.now().month
                start_year = current_year if month_now >= 8 else current_year - 1
                new_schedule = generate_new_year(start_year)
                new_schedule["classes"] = classes
                save_schedule(school_id.strip(), new_schedule)
                st.toast("המוסד נוצר בהצלחה!")
                st.rerun()
            except Exception as ex:
                st.error(f"שגיאה: {ex}")


def _staff_card_html(email: str, role: str, classes: list[str], is_self: bool) -> str:
    """Render a single staff member as an HTML card."""
    role_label = "מנהל" if role == "director" else "מורה"
    chips = "".join(f'<span class="class-chip">{c}</span>' for c in classes)
    self_tag = ' <span style="font-size:0.65rem;color:#9E9E9E;">(את/ה)</span>' if is_self else ""
    return (
        f'<div class="staff-card">'
        f'  <div class="staff-email">{email}{self_tag}</div>'
        f'  <div class="staff-meta">{role_label} &nbsp;|&nbsp; {chips if chips else "ללא כיתות"}</div>'
        f'</div>'
    )


def page_manage_staff(auth_info: dict):
    """Director page: manage teachers and their class permissions.

    Features: card display, class chips, search filter, inline edit,
    email validation, duplicate & self-delete protection.
    """
    school_id = auth_info["school_id"]
    my_email = (auth_info.get("email") or "").lower()
    available_classes = auth_info["allowed_classes"]

    perms = get_permissions(school_id)
    # Build list excluding self
    staff_list = [
        (email, perm) for email, perm in (perms or {}).items()
        if email.lower() != my_email
    ]
    staff_count = len(staff_list)

    # ── Search filter ──
    search_q = st.text_input(
        "חיפוש", key="staff_search",
        placeholder="אימייל...",
    )
    if search_q.strip():
        q = search_q.strip().lower()
        staff_list = [(e, p) for e, p in staff_list if q in e.lower()]

    # ── Staff list ──
    if not staff_list and not search_q.strip():
        st.markdown(
            '<div class="staff-empty">אין מורים רשומים עדיין</div>',
            unsafe_allow_html=True,
        )
    elif not staff_list and search_q.strip():
        st.markdown(
            '<div class="staff-empty">לא נמצאו תוצאות</div>',
            unsafe_allow_html=True,
        )
    else:
        for email, perm in staff_list:
            role = perm.get("role", "teacher")
            classes = perm.get("allowed_classes", [])

            st.markdown(
                _staff_card_html(email, role, classes, is_self=False),
                unsafe_allow_html=True,
            )

            # Action row: Edit / Remove
            act_col1, act_col2 = st.columns(2)
            with act_col1:
                edit_key = f"edit_toggle_{email}"
                if st.button("עריכה", key=f"edit_btn_{email}", use_container_width=True):
                    st.session_state[edit_key] = not st.session_state.get(edit_key, False)

            with act_col2:
                if role == "director":
                    st.button("הסר", key=f"rm_{email}", disabled=True,
                              use_container_width=True, help="לא ניתן להסיר מנהל")
                else:
                    if st.button("הסר", key=f"rm_{email}", use_container_width=True):
                        remove_teacher_permission(school_id, email)
                        st.toast(f"{email} הוסר מהצוות")
                        st.rerun()

            # Inline edit panel
            if st.session_state.get(f"edit_toggle_{email}", False):
                current_role = role  # "teacher" or "director"
                new_role_label = st.radio("תפקיד", ["מורה", "מנהל"], 
                                          index=0 if current_role != "director" else 1,
                                          key=f"edit_role_{email}", horizontal=False)
                 
                new_classes = st.multiselect(
                    "עדכן כיתות",
                    available_classes,
                    default=[c for c in classes if c in available_classes],
                    key=f"edit_classes_{email}",
                    placeholder="בחר כיתות",
                )
                if st.button("שמור שינויים", key=f"save_edit_{email}",
                             type="primary", use_container_width=True):
                    # Validate selection
                    is_director = (new_role_label == "מנהל")
                    if new_classes or is_director: 
                        # Allow director with no specific classes if we want, but keeping safe for now
                        final_role = "director" if is_director else "teacher"
                        set_teacher_permission(school_id, email, new_classes, role=final_role)
                        st.session_state[f"edit_toggle_{email}"] = False
                        st.toast(f"עודכן בהצלחה: {email} ({new_role_label})")
                        st.rerun()
                    else:
                        st.warning("יש לבחור לפחות כיתה אחת (עבור מורה)")

    # ── Add new teacher ──
    st.markdown("---")
    st.markdown(
        '<div class="sb-section-hdr">'
        '<span class="sb-icon">&#10133;</span> הוספת מורה</div>',
        unsafe_allow_html=True,
    )
    with st.form("add_teacher_form"):
        teacher_email = st.text_input("אימייל", placeholder="teacher@school.org")
        role_type = st.radio("תפקיד", ["מורה", "מנהל"], horizontal=False)
        selected_classes = st.multiselect("כיתות", available_classes, placeholder="בחר כיתות")

        if st.form_submit_button("הוסף", type="primary", use_container_width=True):
            email_clean = teacher_email.strip().lower()
            if not email_clean:
                st.warning("הזן אימייל")
            elif not _is_valid_email(email_clean):
                st.error("אימייל לא תקין")
            elif not selected_classes:
                st.warning("בחר לפחות כיתה אחת")
            elif email_clean == my_email:
                st.error("לא ניתן להוסיף את עצמך")
            elif perms and email_clean in perms:
                st.warning(f"{email_clean} כבר בצוות")
            else:
                final_role = "director" if role_type == "מנהל" else "teacher"
                set_teacher_permission(school_id, email_clean, selected_classes, role=final_role)
                st.toast(f"{email_clean} נוסף!")
                st.rerun()

    if staff_count > 0:
        st.caption(f"{staff_count} חברי צוות רשומים")


# ===================================================================
# SIDEBAR SECTIONS
# ===================================================================

def _email_initial(email: str) -> str:
    """Extract first Hebrew/Latin letter for avatar."""
    name = email.split("@")[0] if "@" in email else email
    return name[0].upper() if name else "?"


def _is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email.strip()))


def _top_bar_html(auth_info: dict) -> str:
    """Build the HTML for the top-bar user badge (right-aligned in RTL)."""
    name = auth_info.get("name", "")
    email = auth_info.get("email", "")
    initial = _email_initial(name or email) if email else "?"
    role = auth_info.get("role", "")
    role_label = "מנהל" if role == "director" else "מורה" if role == "teacher" else "צופה"
    role_cls = "director" if role == "director" else "teacher"
    return (
        f'<div class="top-bar-user">'
        f'  <div class="top-bar-avatar">{initial}</div>'
        f'  <div>'
        f'    <div class="top-bar-name">{name}</div>'
        f'    <div class="top-bar-email">{email}</div>'
        f'    <span class="top-bar-role {role_cls}">{role_label}</span>'
        f'  </div>'
        f'</div>'
    )


def render_admin_tab(data: dict, cls: str, school_id: str, auth_info: dict):
    """Full admin tab ג€” expanders inside a styled panel."""
    c1, c2 = st.columns(2)

    with c1:
        with st.container():
            st.markdown(
                '<div class="sb-section-hdr">'
                '<span class="sb-icon">&#9881;</span> ניהול לוח</div>',
                unsafe_allow_html=True,
            )

            with st.expander("&#128218;  הוספת כיתה"):
                nc = st.text_input("שם כיתה חדשה", key="nc", placeholder="יא 4")
                if st.button("הוסף", key="add_class_btn") and nc.strip() and nc.strip() not in data.get("classes", []):
                    add_class_to_school(school_id, nc.strip())
                    data["classes"].append(nc.strip())
                    save_schedule(school_id, data)
                    st.rerun()

            with st.expander("&#128197;  הוספת אירוע חדש"):
                _sidebar_add_event_form(data, cls, school_id, auth_info)

            with st.expander("&#127979;  סנכרון משרד החינוך"):
                _sidebar_ministry_tools(data, cls, school_id)

            with st.expander("&#127796;  ייבוא חופשות וחגים"):
                _sidebar_holidays_import(data, school_id)

            with st.expander("&#128260;  מעבר לשנה חדשה"):
                _sidebar_year_rollover(data, cls, school_id)

            with st.expander("&#128279;  קישור שיתוף להורים"):
                share_class = st.selectbox("כיתה לשיתוף", data.get("classes", []), key="share_class_select")
                detected_url = _get_base_url()
                share_url = (
                    f"{detected_url}?school_id={urllib.parse.quote(school_id)}"
                    f"&class={urllib.parse.quote(share_class)}&mode=view"
                )
                st.code(share_url, language=None)
                st.caption("שלח להורים ג€” צפייה ללא התחברות")

    with c2:
        st.markdown(
            '<div class="sb-section-hdr">'
            '<span class="sb-icon">&#128101;</span> ניהול צוות</div>',
            unsafe_allow_html=True,
        )
        page_manage_staff(auth_info)


def _sidebar_add_event_form(data: dict, cls: str, school_id: str, auth_info: dict):
    """Quick form to add an event."""
    event_text = st.text_input("שם האירוע", key="sidebar_event_text", placeholder="לדוגמה: מבחן")
    col_date, col_type = st.columns(2)
    with col_date:
        event_date = st.date_input("תאריך", key="sidebar_event_date")
    with col_type:
        event_type = st.selectbox(
            "סוג אירוע",
            list(STYLES.keys()),
            format_func=lambda x: STYLES[x]["label"],
            key="sidebar_event_type",
        )
    event_cls = st.selectbox(
        "כיתה יעד",
        auth_info["allowed_classes"] + ["all"],
        key="sidebar_event_cls",
    )
    bagrut_start = ""
    bagrut_end = ""
    if event_type == "bagrut":
        t1, t2 = st.columns(2)
        with t1:
            bagrut_start = st.text_input("שעת התחלה", key="sidebar_bagrut_start", placeholder="09:00")
        with t2:
            bagrut_end = st.text_input("שעת סיום", key="sidebar_bagrut_end", placeholder="12:00")
    if st.button("הוסף", key="sidebar_add_event_btn", type="primary", use_container_width=True):
        if event_text.strip():
            target = datetime.combine(event_date, datetime.min.time())
            loc = date_to_week_day(data["weeks"], target)
            if loc:
                wi, dk = loc
                event_payload = {"text": event_text.strip(), "type": event_type, "class": event_cls}
                if event_type == "bagrut":
                    st_time = _normalize_exam_time(bagrut_start)
                    en_time = _normalize_exam_time(bagrut_end)
                    if st_time and en_time:
                        event_payload["text"] = f"{event_text.strip()} {st_time}-{en_time}"
                    elif st_time:
                        event_payload["text"] = f"{event_text.strip()} {st_time}"
                    event_payload["start_time"] = st_time
                    event_payload["end_time"] = en_time
                data["weeks"][wi]["days"].setdefault(dk, []).append(event_payload)
                save_schedule(school_id, data)
                st.toast("אירוע נוסף!")
                st.rerun()
            else:
                st.warning("התאריך מחוץ לטווח הלוח")
        else:
            st.warning("הזן שם אירוע")


def _sidebar_ministry_tools(data: dict, cls: str, school_id: str):
    """Ministry of Education import + sync tools."""
    meta = get_ministry_meta()
    moed_info = meta.get("moed", "")
    exam_count = meta.get("count", 0)
    if moed_info:
        st.caption(f"{moed_info} | {exam_count} בחינות")

    col_refresh, col_sync = st.columns(2)
    with col_refresh:
        if st.button("עדכן מהמשרד", key="refresh_ministry", use_container_width=True):
            try:
                with st.spinner("מוריד..."):
                    count = refresh_ministry_db_from_web()
                st.toast(f"עודכן! {count} בחינות")
                st.rerun()
            except Exception as ex:
                st.error(f"שגיאה: {ex}")
    with col_sync:
        if st.button("סנכרן תאריכים", key="resync_btn", use_container_width=True):
            changes = resync_dates_with_ministry(data, cls)
            if not changes:
                st.success("הכל מעודכן!")
            else:
                save_schedule(school_id, data)
                for ch in changes:
                    st.markdown(f"- **{ch['name']}** ({ch['code']}): {ch['old_date']} -> {ch['new_date']}")
                    if ch["conflict"]:
                        st.error(ch["conflict"])
                st.rerun()

    search_query = st.text_input(
        "חיפוש",
        key="ministry_text_search",
        placeholder="מקצוע / סמל...",
    )
    if search_query.strip():
        results = search_ministry_exams(search_query)
        if results:
            st.caption(f"{len(results)} תוצאות")
            for exam in results:
                st.markdown(exam_card_html(exam), unsafe_allow_html=True)
                if st.button(f"ייבא", key=f"import_search_{exam['code']}", use_container_width=True):
                    success, msg = import_exam_to_schedule(data, exam, cls)
                    if success:
                        save_schedule(school_id, data)
                        st.toast(msg if msg else f"יובא: {exam['name']}")
                        st.rerun()
                    else:
                        st.info(msg)
        else:
            st.caption("לא נמצאו תוצאות")
    else:
        all_exams = get_ministry_exams()
        all_exams = [e for e in all_exams if e.get("code") != "_metadata"]
        if all_exams:
            exam_options = ["בחר מקצוע..."] + [
                f"{ex['code']} - {ex['name']}" for ex in all_exams
            ]
            selected_option = st.selectbox("מקצוע", exam_options, key="ministry_select", label_visibility="collapsed")
            if selected_option != "בחר מקצוע...":
                sel_code = selected_option.split(" - ")[0].strip()
                exam = get_ministry_exam(sel_code)
                if exam:
                    st.markdown(exam_card_html(exam), unsafe_allow_html=True)
                    if st.button("ייבא ללוח", key=f"import_{exam['code']}", type="primary", use_container_width=True):
                        success, msg = import_exam_to_schedule(data, exam, cls)
                        if success:
                            save_schedule(school_id, data)
                            st.toast(msg if msg else f"יובא: {exam['name']}")
                            st.rerun()
                        else:
                            st.info(msg)


def _run_holidays_import(data: dict, school_id: str):
    """Import holidays/vacations immediately and persist results."""
    added_count = 0
    year_keys = set()
    if data["weeks"]:
        try:
            sy = datetime.strptime(data["weeks"][0]["start_date"], "%Y-%m-%d").year
            year_keys.add(str(sy))
            year_keys.add(str(sy - 1))
        except Exception:
            pass
        try:
            ey = datetime.strptime(data["weeks"][-1]["start_date"], "%Y-%m-%d").year
            year_keys.add(str(ey))
        except Exception:
            pass
    for yk in year_keys:
        holidays_data = get_holidays(yk)
        if not holidays_data:
            continue
        for h in holidays_data.get("holidays", []):
            try:
                hdate = datetime.strptime(h["date"], "%Y-%m-%d")
            except Exception:
                continue
            loc = date_to_week_day(data["weeks"], hdate)
            if loc is None:
                continue
            wi, dk = loc
            cell = data["weeks"][wi]["days"].get(dk, [])
            if any(e["text"] == h["text"] for e in cell):
                continue
            cell.append({"text": h["text"], "type": h.get("type", "holiday"), "class": "all"})
            data["weeks"][wi]["days"][dk] = cell
            added_count += 1
        for v in holidays_data.get("school_vacations", []):
            try:
                vs = datetime.strptime(v["start"], "%Y-%m-%d")
                ve = datetime.strptime(v["end"], "%Y-%m-%d")
            except Exception:
                continue
            d = vs
            while d <= ve:
                loc = date_to_week_day(data["weeks"], d)
                if loc:
                    wi, dk = loc
                    cell = data["weeks"][wi]["days"].get(dk, [])
                    if not any(e["text"] == v["text"] for e in cell):
                        cell.append({"text": v["text"], "type": "vacation", "class": "all"})
                        data["weeks"][wi]["days"][dk] = cell
                        added_count += 1
                d += timedelta(days=1)
    if added_count > 0:
        save_schedule(school_id, data)
        st.toast(f"יובאו {added_count} אירועים!")
        st.rerun()
    else:
        st.info("הכל כבר קיים בלוח")


def _sidebar_holidays_import(data: dict, school_id: str):
    """Import holidays and vacations."""
    st.caption("ייבוא חגים, חופשות וימי זיכרון לכל השנה")
    if st.button("ייבא עכשיו", key="import_holidays_btn", type="primary", use_container_width=True):
        _run_holidays_import(data, school_id)


def _sidebar_year_rollover(data: dict, cls: str, school_id: str):
    """Generate a new academic year."""
    st.caption("לוח חדש עם חגים, חופשות ופרשות")
    current_year = datetime.now().year
    col_yr, col_bg = st.columns([1, 1])
    with col_yr:
        new_year_start = st.number_input(
            "שנה", min_value=2024, max_value=2040,
            value=current_year, key="new_year_input",
        )
    with col_bg:
        st.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
        import_bagrut = st.checkbox("כולל בגרויות", value=True, key="import_bagrut_check")
    if st.button("צור שנה חדשה", key="gen_new_year", type="primary", use_container_width=True):
        with st.spinner("יוצר לוח שנה..."):
            new_data = generate_new_year(int(new_year_start))
            new_data["classes"] = data["classes"]

            if import_bagrut:
                all_exams = get_ministry_exams()
                all_exams = [e for e in all_exams if e.get("code") != "_metadata"]
                db_base_year = None
                for ex in all_exams:
                    try:
                        db_base_year = datetime.strptime(ex["date"], "%Y-%m-%d").year
                        break
                    except Exception:
                        continue
                if db_base_year:
                    target_exam_year = int(new_year_start) + 1
                    year_offset = target_exam_year - db_base_year
                else:
                    year_offset = 0

                for exam in all_exams:
                    try:
                        orig_date = datetime.strptime(exam["date"], "%Y-%m-%d")
                        shifted_date = orig_date.replace(year=orig_date.year + year_offset) if year_offset else orig_date
                        loc = date_to_week_day(new_data["weeks"], shifted_date)
                        if loc:
                            wi, dk = loc
                            label = _build_bagrut_label(exam)
                            new_data["weeks"][wi]["days"][dk].append({
                                "text": label,
                                "type": "bagrut",
                                "class": cls,
                                "exam_code": exam["code"],
                                "start_time": _normalize_exam_time(exam.get("start_time")),
                                "end_time": _normalize_exam_time(exam.get("end_time")),
                            })
                    except Exception:
                        continue

            save_schedule(school_id, new_data)
        st.toast("לוח שנה חדש נוצר!")
        st.rerun()


def render_export_tab(data: dict, cls: str, filtered_weeks: list):
    """Export & Share tab ג€” horizontal grid layout."""
    st.markdown(
        '<div class="sb-section-hdr">'
        '<span class="sb-icon">&#128229;</span> ייצוא ושיתוף</div>',
        unsafe_allow_html=True,
    )

    col_xl, col_wa, col_png = st.columns(3)

    with col_xl:
        # Cache Excel bytes so they aren't rebuilt on every rerun
        xl_cache_key = _export_cache_key(data, cls, filtered_weeks)
        if st.session_state.get("xl_cache_key") != xl_cache_key:
            st.session_state["xl_cache_key"] = xl_cache_key
            st.session_state["xl_bytes"] = to_excel(data, cls)
        st.download_button(
            "&#128230;  הורד Excel",
            data=st.session_state["xl_bytes"],
            file_name=f"לוח_{cls}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
            use_container_width=True,
        )

    with col_wa:
        wa_text = build_whatsapp_text(data, cls, filtered_weeks)
        wa_url = f"https://wa.me/?text={urllib.parse.quote(wa_text[:4000])}"
        st.markdown(
            f'<a href="{wa_url}" target="_blank" class="sb-export-link" '
            f'style="background:#25D366;color:#fff;">'
            f'&#128242;  שתף בווצאפ</a>',
            unsafe_allow_html=True,
        )

    with col_png:
        if "wa_png_cache_key" not in st.session_state:
            st.session_state["wa_png_cache_key"] = None
            st.session_state["wa_png_bytes"] = None

        cache_key = _export_cache_key(data, cls, filtered_weeks)
        png_image = st.session_state["wa_png_bytes"]
        png_ready = (st.session_state["wa_png_cache_key"] == cache_key and png_image is not None)

        if png_ready:
            is_png = isinstance(png_image, bytes) and len(png_image) > 4 and png_image[:4] == b'\x89PNG'
            if is_png:
                st.download_button(
                    "&#128248;  הורד תמונה PNG",
                    data=png_image,
                    file_name=f"לוח_{cls}.png",
                    mime="image/png",
                    key="download_png_tab",
                    use_container_width=True,
                )
            else:
                st.download_button(
                    "&#128248;  הורד תמונה (HTML)",
                    data=png_image,
                    file_name=f"לוח_{cls}.html",
                    mime="text/html",
                    key="download_html_tab",
                    use_container_width=True,
                )
        else:
            if st.button("&#128248;  צור תמונה PNG", key="gen_png_tab", use_container_width=True):
                with st.spinner("יוצר תמונה..."):
                    png_result = schedule_to_png(data, cls, filtered_weeks)
                st.session_state["wa_png_cache_key"] = cache_key
                st.session_state["wa_png_bytes"] = png_result
                st.rerun()


# ===================================================================
# DATE RANGE FILTER HELPER
# ===================================================================

def _compute_default_date_range(data: dict):
    """Return default range = current week start (Sunday) .. end of academic August."""
    today = datetime.now().date()
    # Week starts on Sunday (aligned with DAY_KEYS: sunday..shabbat)
    days_since_sunday = (today.weekday() + 1) % 7
    start = today - timedelta(days=days_since_sunday)

    # Academic end target: August 31.
    # Jan-Aug -> current year, Sep-Dec -> next year.
    august_year = today.year if today.month <= 8 else today.year + 1
    end = datetime(august_year, 8, 31).date()

    # Clamp to schedule range
    if data["weeks"]:
        try:
            sched_start = datetime.strptime(data["weeks"][0]["start_date"], "%Y-%m-%d").date()
            sched_end = datetime.strptime(data["weeks"][-1]["start_date"], "%Y-%m-%d").date() + timedelta(days=6)
            start = max(start, sched_start)
            end = min(end, sched_end)
            if start > end:
                start, end = sched_start, sched_end
        except Exception:
            pass
    return start, end


def _filter_weeks_by_range(data: dict, range_start, range_end):
    """Return list of (index, week) tuples whose date span overlaps with [range_start, range_end]."""
    filtered = []
    for wi, wk in enumerate(data["weeks"]):
        try:
            sd = datetime.strptime(wk["start_date"], "%Y-%m-%d").date()
            ed = sd + timedelta(days=6)
            if ed >= range_start and sd <= range_end:
                filtered.append((wi, wk))
        except Exception:
            filtered.append((wi, wk))
    return filtered


# ===================================================================
# MAIN SCHEDULER VIEW
# ===================================================================

@st.fragment
def render_scheduler(data: dict, cls: str, auth_info: dict, filtered_weeks: list):
    """Render the main visual schedule grid."""
    school_id = auth_info.get("school_id", "")
    is_director = auth_info["role"] == "director"
    is_teacher = auth_info["role"] == "teacher"
    can_edit = is_director or is_teacher

    # ── Header ──
    year_label = data.get("year", "")
    school_name = auth_info.get("school_name", "")

    st.markdown(
        f'<div class="app-title">'
        f'<h1>לוח מבחנים {year_label}</h1>'
        f'<p>2025 - 2026 &nbsp;|&nbsp; {school_name} &nbsp;|&nbsp; {cls}</p>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Legend ──
    lg = "".join(
        f'<span class="legend-chip" style="background:{s["bg"]};color:{s["fg"]};'
        f'font-weight:{"700" if s["bold"] else "400"};">{s["label"]}</span>'
        for s in STYLES.values()
    )
    lg += '<span class="legend-chip cal-parasha">פרשת השבוע</span>'
    st.markdown(f'<div class="legend-row">{lg}</div>', unsafe_allow_html=True)

    pm = data.get("parashat_hashavua", {})

    # ── Build entire table as single HTML block (much faster than st.columns per row) ──
    table_html = (
        '<table class="sched-table">'
        '<thead><tr>'
    )
    for dn in DAY_NAMES:
        table_html += f'<th class="cal-hdr">{dn}</th>'
    table_html += '</tr></thead><tbody>'

    for wi, wk in filtered_weeks:
        parasha = pm.get(wk["start_date"], "")
        even = wi % 2 == 0
        table_html += '<tr>'
        for di, dk in enumerate(DAY_KEYS):
            day_date = get_day_date(wk.get("start_date", ""), di)
            evs = [e for e in wk["days"].get(dk, []) if e.get("class") in (cls, "all")]
            chips = "".join(chip_html(e) for e in evs)
            if dk == "shabbat" and parasha:
                chips += f'<span class="cal-parasha">{parasha}</span>'
            bg = "#F8F9FA" if even else "#FFFFFF"
            table_html += (
                f'<td style="background:{bg};border:1px solid #DEE2E6;'
                f'padding:4px 3px;min-height:72px;text-align:center;vertical-align:top;">'
                f'<div style="color:#90A4AE;font-size:0.7em;font-weight:500;">{day_date}</div>'
                f'{chips}</td>'
            )
        table_html += '</tr>'
    table_html += '</tbody></table>'
    st.markdown(table_html, unsafe_allow_html=True)

    # ── Edit buttons row (only for editors) ──
    if can_edit:
        if "edit_cell" not in st.session_state:
            st.session_state["edit_cell"] = None

        # Render "+" buttons in columns per week — only buttons, no heavy widgets
        for wi, wk in filtered_weeks:
            rcols = st.columns(7)
            for di, dk in enumerate(DAY_KEYS):
                with rcols[di]:
                    is_selected = st.session_state["edit_cell"] == (wi, di)
                    if st.button(
                        "✏️" if is_selected else "+",
                        key=f"sel_{wi}_{di}",
                        use_container_width=True,
                        type="primary" if is_selected else "secondary",
                    ):
                        if is_selected:
                            st.session_state["edit_cell"] = None
                        else:
                            st.session_state["edit_cell"] = (wi, di)
                        st.rerun()

        # ── Single popover-style edit form for the selected cell ──
        if st.session_state["edit_cell"] is not None:
            sel_wi, sel_di = st.session_state["edit_cell"]
            sel_dk = DAY_KEYS[sel_di]
            sel_wk = None
            for fwi, fwk in filtered_weeks:
                if fwi == sel_wi:
                    sel_wk = fwk
                    break
            if sel_wk is None:
                st.session_state["edit_cell"] = None
            else:
                sel_date = get_day_date(sel_wk.get("start_date", ""), sel_di)
                all_cell = sel_wk["days"].get(sel_dk, [])
                vis = [e for e in all_cell if e.get("class") in (cls, "all")]

                with st.popover(f"✏️ {DAY_NAMES[sel_di]} {sel_date}", use_container_width=True):
                    for idx_ev, ev in enumerate(vis):
                        if st.button(
                            f"מחק {ev['text']}", key=f"del_{sel_wi}_{sel_di}_{idx_ev}",
                            use_container_width=True,
                        ):
                            all_cell.remove(ev)
                            save_schedule(school_id, data)
                            st.session_state["edit_cell"] = None
                            st.rerun()

                    nt = st.text_input("שם האירוע", key="edit_ev_name", placeholder="לדוגמה: מבחן")
                    tp = st.selectbox(
                        "סוג אירוע", list(STYLES.keys()),
                        format_func=lambda x: STYLES[x]["label"],
                        key="edit_ev_type",
                    )
                    ecls = st.selectbox(
                        "כיתה יעד", auth_info["allowed_classes"] + ["all"],
                        key="edit_ev_cls",
                    )
                    bagrut_start = ""
                    bagrut_end = ""
                    if tp == "bagrut":
                        tcol1, tcol2 = st.columns(2)
                        with tcol1:
                            bagrut_start = st.text_input(
                                "שעת התחלה",
                                key="edit_ev_start",
                                placeholder="09:00",
                            )
                        with tcol2:
                            bagrut_end = st.text_input(
                                "שעת סיום",
                                key="edit_ev_end",
                                placeholder="12:00",
                            )
                    if st.button("הוסף", key="edit_ev_add", type="primary", use_container_width=True):
                        if nt.strip():
                            if sel_dk not in sel_wk["days"]:
                                sel_wk["days"][sel_dk] = []
                            event_payload = {"text": nt.strip(), "type": tp, "class": ecls}
                            if tp == "bagrut":
                                st_time = _normalize_exam_time(bagrut_start)
                                en_time = _normalize_exam_time(bagrut_end)
                                if st_time and en_time:
                                    event_payload["text"] = f"{nt.strip()} {st_time}-{en_time}"
                                elif st_time:
                                    event_payload["text"] = f"{nt.strip()} {st_time}"
                                event_payload["start_time"] = st_time
                                event_payload["end_time"] = en_time
                            sel_wk["days"][sel_dk].append(event_payload)
                            save_schedule(school_id, data)
                            st.session_state["edit_cell"] = None
                            st.rerun()


# ===================================================================
# PUBLIC LINK HELPER
# ===================================================================

def _get_base_url() -> str:
    try:
        ctx = st.context
        if hasattr(ctx, "headers"):
            host = ctx.headers.get("Host", "")
            scheme = ctx.headers.get("X-Forwarded-Proto", "https")
            if host:
                return f"{scheme}://{host}"
    except Exception:
        pass
    return "http://localhost:8501"


# ===================================================================
# MAIN
# ===================================================================

def main():
    st.set_page_config(page_title="לוח מבחנים", layout="wide", initial_sidebar_state="collapsed")
    st.markdown(APP_CSS, unsafe_allow_html=True)

    # ── Authenticate ──
    try:
        auth_info = authenticate()
    except Exception as e:
        import traceback
        st.error(f"Critical Auth Error: {e}")
        st.text(traceback.format_exc())
        return

    if not auth_info["authenticated"]:
        return

    # ── Public mode: read-only, no tabs ──
    if auth_info["is_public"]:
        school_id = auth_info["school_id"]
        if not school_id:
            st.error("קישור לא תקין - חסר מזהה מוסד")
            return
        data = get_schedule(school_id)
        if not data["weeks"]:
            st.error("לא נמצאו נתונים למוסד זה")
            return
        allowed = auth_info["allowed_classes"]
        cls = allowed[0] if allowed else (data["classes"][0] if data["classes"] else "יא 1")

        default_start, default_end = _compute_default_date_range(data)
        filtered_weeks = _filter_weeks_by_range(data, default_start, default_end)

        render_scheduler(data, cls, auth_info, filtered_weeks)
        return

    # ── No school yet → create school flow ──
    if not auth_info["school_id"] and not auth_info["schools"]:
        page_create_school(auth_info)
        return

    if not auth_info["school_id"]:
        st.info("בחר מוסד מהרשימה")
        return

    # ── Load schedule data ──
    school_id = auth_info["school_id"]
    data = get_schedule(school_id)
    is_director = auth_info["role"] == "director"
    is_teacher = auth_info["role"] == "teacher"
    allowed_classes = auth_info["allowed_classes"]

    # ──────────────────────────────────────────
    # TOP BAR ג€” user info + class selector + date range
    # ──────────────────────────────────────────
    year_label = data.get("year", "")
    school_name = auth_info.get("school_name", "")

    user_html = _top_bar_html(auth_info)
    title_html = (
        f'<div class="top-bar-title">'
        f'  <h2>לוח מבחנים {year_label}</h2>'
        f'  <p>2025 - 2026 &nbsp;|&nbsp; {school_name}</p>'
        f'</div>'
    )

    bar_col, logout_col = st.columns([7, 1])
    with bar_col:
        st.markdown(
            f'<div class="top-bar">{user_html}{title_html}</div>',
            unsafe_allow_html=True,
        )
    with logout_col:
        st.markdown('<div class="logout-btn" style="padding-top:16px">', unsafe_allow_html=True)
        if st.button("יציאה", key="logout_btn", use_container_width=True):
            for k in ("auth_email", "auth_name", "auth_token", "auth_uid", "auth_token_exp", "selected_school_id"):
                st.session_state.pop(k, None)
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # Shared selector defaults (controls are rendered above the table in main column)
    class_options = allowed_classes if allowed_classes else data.get("classes", ["יא 1"])
    default_start, default_end = _compute_default_date_range(data)
    panel_cls = st.session_state.get("class_select", class_options[0] if class_options else "יא 1")

    # ──────────────────────────────────────────
    # LAYOUT: side buttons + main schedule
    # ──────────────────────────────────────────
    if "open_panel" not in st.session_state:
        st.session_state["open_panel"] = None

    # Build button definitions based on role
    action_buttons = []
    if is_director:
        action_buttons += [
            ("add_event", "\U0001F4CC הוספת אירוע", "panel"),
            ("ministry", "\U0001F393 בגרויות", "panel"),
            ("holidays", "\U0001F334 חופשות וחגים", "action"),
            ("new_year", "\U0001F5D3 שנה חדשה", "panel"),
            ("add_class", "\U0001F3EB כיתה חדשה", "panel"),
            ("staff", "\U0001F465 ניהול צוות", "panel"),
        ]

    if action_buttons:
        col_main, col_side = st.columns([4, 1])
    else:
        col_main = st.container()
        col_side = None

    # ── Side buttons column ──
    if col_side is not None:
        with col_side:
            with st.container():
                st.markdown('<div class="side-actions-marker"></div>', unsafe_allow_html=True)
                st.markdown('<div class="side-actions-title">ניהול מהיר לפי צבע ואייקון</div>', unsafe_allow_html=True)
                for key, label, action_kind in action_buttons:
                    is_open = st.session_state["open_panel"] == key
                    btn_type = "primary" if action_kind == "panel" and is_open else "secondary"
                    if st.button(label, key=f"nav_{key}", use_container_width=True, type=btn_type):
                        if action_kind == "action":
                            st.session_state["open_panel"] = None
                            _run_holidays_import(data, school_id)
                        else:
                            st.session_state["open_panel"] = None if is_open else key
                            st.rerun()

            # ── Render open panel ──
            panel = st.session_state["open_panel"]

            if panel == "add_event" and is_director:
                with st.container():
                    st.markdown('<div class="side-panel-marker"></div>', unsafe_allow_html=True)
                    st.markdown('<div class="side-panel-heading">הוספת אירוע</div>', unsafe_allow_html=True)
                    _sidebar_add_event_form(data, panel_cls, school_id, auth_info)

            elif panel == "ministry" and is_director:
                with st.container():
                    st.markdown('<div class="side-panel-marker"></div>', unsafe_allow_html=True)
                    st.markdown('<div class="side-panel-heading">סנכרון בגרויות</div>', unsafe_allow_html=True)
                    _sidebar_ministry_tools(data, panel_cls, school_id)

            elif panel == "new_year" and is_director:
                with st.container():
                    st.markdown('<div class="side-panel-marker"></div>', unsafe_allow_html=True)
                    st.markdown('<div class="side-panel-heading">מעבר לשנה חדשה</div>', unsafe_allow_html=True)
                    _sidebar_year_rollover(data, panel_cls, school_id)

            elif panel == "add_class" and is_director:
                with st.container():
                    st.markdown('<div class="side-panel-marker"></div>', unsafe_allow_html=True)
                    st.markdown('<div class="side-panel-heading">כיתה חדשה</div>', unsafe_allow_html=True)
                    nc = st.text_input("שם", key="nc", placeholder="יא 4")
                    if st.button("הוסף כיתה", key="add_class_btn", type="primary", use_container_width=True):
                        if nc.strip() and nc.strip() not in data.get("classes", []):
                            add_class_to_school(school_id, nc.strip())
                            data["classes"].append(nc.strip())
                            save_schedule(school_id, data)
                            st.toast(f"כיתה '{nc.strip()}' נוספה!")
                            st.rerun()

            elif panel == "staff" and is_director:
                with st.container():
                    st.markdown('<div class="side-panel-marker"></div>', unsafe_allow_html=True)
                    st.markdown('<div class="side-panel-heading">ניהול צוות</div>', unsafe_allow_html=True)
                    page_manage_staff(auth_info)

    # ── Main column: export row + schedule ──
    with col_main:
        # ── Controls row: class + date range (sticky above table) ──
        with st.container():
            st.markdown('<div class="ctrl-row-marker"></div>', unsafe_allow_html=True)
            ctrl1, ctrl2 = st.columns([1.2, 2.2])
            with ctrl1:
                cls = st.selectbox("כיתה", class_options, key="class_select", label_visibility="collapsed")
            with ctrl2:
                date_range = st.date_input(
                    "תאריכים",
                    value=(default_start, default_end),
                    key="date_range_filter",
                    label_visibility="collapsed",
                )

        if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
            range_start, range_end = date_range
        elif isinstance(date_range, (list, tuple)) and len(date_range) == 1:
            range_start = date_range[0]
            range_end = default_end
        else:
            range_start = default_start
            range_end = default_end

        filtered_weeks = _filter_weeks_by_range(data, range_start, range_end)

        # ── Compact export toolbar above the table ──
        if is_director or is_teacher:
          with st.container():
            st.markdown('<div class="export-bar-marker"></div>', unsafe_allow_html=True)
            exp_c1, exp_c2, exp_c3 = st.columns([1, 1, 1])

            # 1) Copy share link (using popover with st.code for native copy)
            with exp_c1:
                with st.popover("\U0001F517 העתק קישור", use_container_width=True):
                    detected_url = _get_base_url()
                    share_url = f"{detected_url}?school_id={urllib.parse.quote(school_id)}&class={urllib.parse.quote(cls)}&mode=view"
                    st.caption("לחץ על הכפתור להעתקה:")
                    st.code(share_url, language=None)

            # 2) Download Excel
            with exp_c2:
                xl_cache_key = _export_cache_key(data, cls, filtered_weeks)
                if st.session_state.get("xl_cache_key") != xl_cache_key:
                    st.session_state["xl_cache_key"] = xl_cache_key
                    st.session_state["xl_bytes"] = to_excel(data, cls)
                st.download_button(
                    "\U0001F4CA הורד Excel",
                    data=st.session_state["xl_bytes"],
                    file_name=f"לוח_{cls}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                    use_container_width=True,
                    key="top_download_excel",
                )

            # 3) Download Image
            with exp_c3:
                if "wa_png_cache_key" not in st.session_state:
                    st.session_state["wa_png_cache_key"] = None
                    st.session_state["wa_png_bytes"] = None

                cache_key = _export_cache_key(data, cls, filtered_weeks)
                png_image = st.session_state["wa_png_bytes"]
                png_ready = (st.session_state["wa_png_cache_key"] == cache_key and png_image is not None)

                if png_ready:
                    is_png = isinstance(png_image, bytes) and len(png_image) > 4 and png_image[:4] == b'\x89PNG'
                    if is_png:
                        st.download_button(
                            "\U0001F5BC\uFE0F הורד תמונה",
                            data=png_image,
                            file_name=f"לוח_{cls}.png",
                            mime="image/png",
                            key="top_download_png",
                            use_container_width=True,
                        )
                    else:
                        st.download_button(
                            "\U0001F5BC\uFE0F הורד תמונה (HTML)",
                            data=png_image,
                            file_name=f"לוח_{cls}.html",
                            mime="text/html",
                            key="top_download_html",
                            use_container_width=True,
                        )
                else:
                    if st.button("\U0001F5BC\uFE0F צור תמונה", key="top_gen_png", use_container_width=True):
                        with st.spinner("\U0001F5BC\uFE0F יוצר תמונה..."):
                            png_result = schedule_to_png(data, cls, filtered_weeks)
                        st.session_state["wa_png_cache_key"] = cache_key
                        st.session_state["wa_png_bytes"] = png_result
                        st.rerun()

        # ── Schedule table ──
        render_scheduler(data, cls, auth_info, filtered_weeks)


if __name__ == "__main__":
    main()
